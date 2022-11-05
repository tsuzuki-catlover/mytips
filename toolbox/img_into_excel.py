import openpyxl


def add_image_to_ws(ws, img_path, col_idx, row_idx, scale=1):
    img = openpyxl.drawing.image.Image(img_path)
    img.height = img.height * scale
    img.width = img.width * scale
    col_str = openpyxl.utils.get_column_letter(col_idx)

    ws.row_dimensions[row_idx].height = 8
    ws.row_dimensions[row_idx].height = \
        max(img.height * 0.75, ws.row_dimensions[row_idx].height)

    ws.column_dimensions[col_str].width = 15
    ws.column_dimensions[col_str].width = \
        max(img.width * 0.128, ws.column_dimensions[col_str].width)

    cell_position = ws.cell(row=row_idx, column=col_idx).coordinate
    img.anchor = cell_position
    ws.add_image(img)


if __name__ == '__main__':
    output = 'excel_with_img.xlsx'
    img_path = 'examples/lenna.png'
    col = 2
    row = 2

    # Open file and sheet
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    add_image_to_ws(ws, img_path, col, row, 0.5)

    # Save
    wb.save(output)
